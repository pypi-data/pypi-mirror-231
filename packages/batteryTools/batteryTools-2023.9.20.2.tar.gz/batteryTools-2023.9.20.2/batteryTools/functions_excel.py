__update__ = '2023.09.01'
__author__ = 'PABLO GONZALEZ PILA <pablogonzalezpila@gmail.com>'

'''
NOTES:
    Conjunto de librerias para uso de excel
TASK:
    Ver actualizaciones de openpyxl para RichText
    Usar el en los metodos self.ROW
    Añadir la funcion de trabajar con letras y enteros:
        ejm: [ COL = xls.utils.get_column_letter(COLUMN) ]
    Añadir libreria xlsxwriter para uso de hojas EXCEL ya creadas
    ** No me termina de gustar es usar save en los metodos
WARNINGS:
    < xlsxwriter > Instalado para usar texto enriquecido (En pruebas)
'''

''' SYSTEM LIBRARIES '''
import os
import pandas as pd

''' CUSTOM MAIN LIBRARIES '''

''' 
OPENPYXL REPORT
--------------------------------------------------------
'''

import openpyxl as xls
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font
from openpyxl.styles import borders
from openpyxl.styles.borders import Border
from openpyxl.worksheet import pagebreak
from openpyxl.utils import get_column_letter

class XLSREPORT:
    '''
    Excel book and functions \n
    DEBUG:\n
        - En el caso de WorkSheet en __init__ hay que preguntar y no usar try/except
        - Hay que depurar el tratamiento de archivos (repetidos, protegidos, etc)
    '''
    def __init__(self, filePath: str, FONT="Arial", WS_NAME="Data"):
        self.FILE = filePath
        self.FILE_STR = self.FILE + ".xlsx"
        self.fileName = os.path.basename(self.FILE_STR)
        self.ROW = 1
        self.FONT = FONT
        self.WS_NAME = WS_NAME
        ## WORKBOOK
        try:
            self.WB = xls.load_workbook(self.FILE_STR)
            # print("LOAD")
        except:
            self.WB = xls.Workbook(self.FILE_STR)
            self.WB.create_sheet(self.WS_NAME)
            self.WB.save(self.FILE_STR)
            self.WB.close
            self.WB = xls.load_workbook(self.FILE_STR)
        ## WORKSHEET
        try:
            self.WS = self.WB[self.WS_NAME]
        except:
            self.WB.create_sheet(self.WS_NAME)
            self.WS = self.WB[self.WS_NAME]
        # self.WS.dimensions.ColumnDimension(self.WS, bestFit=True)

    def SAVE(self):
        self.WB.save(self.FILE_STR)

    def CLOSE(self):
        self.WB.close

    def SHEET_NEW(self, sheet_name: str):
        '''
        Create a new excel sheet
        '''
        self.WB.create_sheet(sheet_name)

    def SHEET_SELECT(self, sheet_name):
        self.WS = self.WB[sheet_name]

    def WR(self, ROW: int = 1, COLUMN: int = 1, VALUE="", f_size: int = 10, f_bolf=False, save: bool = False):
        # self.WS["A4"].value = VALUE
        self.WS.cell(ROW, COLUMN).value = VALUE
        self.WS.cell(ROW, COLUMN).alignment = Alignment(horizontal='left', vertical='center')
        self.WS.cell(ROW, COLUMN).font = Font(name=self.FONT, size=f_size, bold=f_bolf,)
        if save == True: self.SAVE()
    
    def WR_TITLE(self, ROW: int = 1, COLUMN: int = 1, VALUE="", save: bool = False):
        self.WS.cell(ROW, COLUMN).value = VALUE
        self.WS.cell(ROW, COLUMN).alignment = Alignment(horizontal='left', vertical='center')
        self.WS.cell(ROW, COLUMN).font = Font(name=self.FONT, size=12, bold=True,)
        self.ROW_WIDTH(ROW, 40)
        if save: self.SAVE()

    def WR_HEADER(self, ROW: int = 1, COLUMN: int = 1, VALUE="", VERTICAL: str = "center", wrap: bool = False, save: bool = False):
        '''
        Write in cell with HEADER style font
        Bold, Size 10
        VERTICAL: "top", "center", "bottom", "justify", "distributed"
        '''
        self.WS.cell(ROW, COLUMN).value = VALUE
        self.WS.cell(ROW, COLUMN).alignment = Alignment(horizontal='left', vertical=VERTICAL, wrap_text=wrap)
        self.WS.cell(ROW, COLUMN).font = Font(name=self.FONT, size=10, bold=True)
        if save: 
            self.SAVE()

    def WR_HEADERS(self, ROW: int = 1, HEADERS: list = list, VERTICAL: str = "center", wrap: bool = False, save: bool = False):
        '''
        Write and edit format of Headers List
        '''
        for head in HEADERS:
            self.WR_HEADER(ROW=ROW, COLUMN=HEADERS.index(head)+1, VALUE=head, VERTICAL=VERTICAL, wrap=wrap)
        self.ROW_WIDTH(ROW, 35)
        if save:
            self.SAVE()

    def ROW_INC(self, NUMBER=1):
        '''
        Add an increment in row count
        '''
        self.ROW = self.ROW + int(NUMBER)

    def ROW_GET(self):
        '''
        Get current row
        '''
        FILA = self.ROW
        return int(FILA)

    def ROW_WIDTH(self, ROW=1, WIDTH=10, save: bool = False):
        '''
        Set height of a row
        '''
        self.WS.row_dimensions[ROW].height = WIDTH
        # 
        if save == True: self.SAVE()

    def COL_WIDTH(self, COL: int = 1, WIDTH: int = 20, save=False):
        '''
        '''
        COL_LETT = COLUMN_STR(COL)
        self.WS.column_dimensions[COL_LETT].width = WIDTH
        if save == True: self.SAVE()

    def COL_AUTOFIT(self, save=False):
        '''
        '''
        for column_cells in self.WS.columns:
                new_column_length = max(len(str(cell.value)) for cell in column_cells)
                new_column_letter = (xls.utils.get_column_letter(column_cells[0].column))
                if new_column_length > 0:
                    self.WS.column_dimensions[new_column_letter].width = new_column_length*1.23
        if save == True: self.SAVE()
    
    def COL_FILTERS(self, save: bool =False):
        '''
        Set filters in current WorkSheet from A1 to maximun column and maximun row
        '''
        fullRange = f"A1:{get_column_letter(self.WS.max_column)}{self.WS.max_row}"
        self.WS.auto_filter.ref = fullRange
        if save == True: self.SAVE()

    def LOW_BORDER(self, ROW=1, col_ini=1, col_fin=300, save=False):
        '''
        INCOMPLETE, DEBUG:
        Hay que saber bien el diseño y todas las funciones de borders
        https://openpyxl.readthedocs.io/en/stable/styles.html?highlight=border_style
        '''
        # Style = "thick" (Grueso)
        border0 = borders.Side(style = None, color = None, border_style = None)
        borderLow = borders.Side(
            style = "medium", 
            color="000000", 
            # border_style = "double"
            )
        thin = Border(left = border0, right = border0, bottom = borderLow, top = border0)
        for col in range(col_ini, col_fin): self.WS.cell(row=ROW, column=col).border = thin
        if save == True: self.SAVE()

    # def MERGE(self, ROW_INI, COL_INI, ROW_FIN, COL_FIN, save=False):
    #     self.WS.merge_cells(
    #         start_row = ROW_INI, 
    #         start_column = COL_INI, 
    #         end_row = ROW_FIN, 
    #         end_column = COL_FIN)
    #     if save == True: self.SAVE()

    def PRNT_AREA(self, COL_FIN: int, save: bool = False):
        '''
        Ajusta la zona de impresion
        INCOMPLETE
        '''
        # self.WS = self.WB[SHEET]
        # self.WS.page_setup.orientation = self.WS.ORIENTATION_LANDSCAPE
        self.WS.page_setup.fitToPage = True
        self.WS.page_setup.fitToPage = 1
        self.WS.page_setup.fitToHeight = False
        COL_STR = COLUMN_STR(COL_FIN)
        self.WS.print_area = "A:" + COL_STR
        if save == True: self.SAVE()

    def SHEET_HEAD(self, ROW_FIN: int, save: bool = False):
        '''
        '''
        self.WS.print_title_rows = "1:" + str(ROW_FIN)
        self.WS.page_margins.top = 0.4
        self.WS.page_margins.botom = 0.4
        # self.WS.page_margins.header = 0.7
        self.WS.page_margins.header = 0.4
        self.WS.page_margins.footer = 0.4
        if save == True: self.SAVE()

    def PAGE_BREAK(self, ROW: int = 1, save: bool = False):
        '''
        Insert a page break in selected row
        '''
        page_break = pagebreak.Break(id=ROW-1)
        break_list = self.WS.row_breaks
        break_list.append(page_break)
        self.WS.row_breaks = break_list
        if save == True: self.SAVE()

    def IMAGE_INSERT(self, ROW: int = 1, COLUMN: int = 1, HEIGHT=None, WIDTH=None):
        '''
        * Necesary install Pillow packages
        INCOMPLETE
        Sin usar ni comprobar
        '''
        PATH = "C:/Users/GONZA_PA/Desktop/rus-logo.png"
        img = Image(PATH)
        ## PIXEL VALUE
        if HEIGHT and WIDTH: 
            img.height = HEIGHT
            img.width = WIDTH
        cell_str = xls.utils.get_column_letter(COLUMN) + str(ROW)
        self.WS.add_image(img, cell_str)

def CELL_STR(ROW: int, COLUMN: int) -> str:
    '''
    '''
    col_str = xls.utils.get_column_letter(COLUMN)
    cellstr = "{}{}".format(col_str, ROW)
    return cellstr

def COLUMN_STR(COLUMN: int) -> str:
    '''
    '''
    letter = xls.utils.get_column_letter(COLUMN)
    return letter

def DF_REPORT(DATAFRAME: pd.DataFrame, fileName: str, fontName: str = 'Calibri') -> None:
    '''
    Create excel report from selected Pandas DataFrame
    '''
    report = XLSREPORT(fileName, fontName)
    ## HEADERS
    HEADERS: list = DATAFRAME.columns.values.tolist()
    report.WR_HEADERS(1, HEADERS, VERTICAL="top")
    report.COL_FILTERS()
    report.LOW_BORDER(report.ROW, col_fin=len(HEADERS)+1)
    report.ROW_INC()
    ## DATA
    for row in range(len(DATAFRAME.index)):
        row_data = list(DATAFRAME.iloc[row].values)
        for value in row_data:
            report.WR(report.ROW, row_data.index(value)+1, value)
        report.ROW_INC()
    report.COL_AUTOFIT()
    ## FIN
    report.SAVE()


''' 
--------------------------------------------------------
'''