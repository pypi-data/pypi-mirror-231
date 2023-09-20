import openpyxl
import math

ISBN_KEY = "ISBN-13"
MAX_WIDTH = round(12.94 * 3.5)
MAGIC_BUFFER = 1.23
OUT = "out.xlsx"

# open worksheet
wb = openpyxl.load_workbook("resources/test_sheet.xlsx")
ws = wb["Sheet1"]

# Insert columns
ws.insert_cols(1)
ws.insert_cols(1)
ws.insert_cols(1)
ws.insert_cols(1)
ws["B1"] = "Image"
ws["C1"] = "Order"
ws["D1"] = "Confirmed"

# Find ISBN index
for cell in ws[1]:
    if type(cell.value) is str and cell.value.upper() == ISBN_KEY:
        break

isbn_idx = cell.column

# Move ISBN to column "A"
ws.move_range(
    f"{cell.column_letter}{cell.row}:{cell.column_letter}{ws.max_row}",
    cols=-(cell.column - 1),
)
ws.delete_cols(isbn_idx)

# Create column widths
dim_holder = openpyxl.worksheet.dimensions.DimensionHolder(worksheet=ws)

# Hard code first two columns
dim_holder["A"] = openpyxl.worksheet.dimensions.ColumnDimension(
    ws, index="A", width=math.ceil(13 * MAGIC_BUFFER)
)
dim_holder["B"] = openpyxl.worksheet.dimensions.ColumnDimension(
    ws, index="B", width=18.0
)

# Dynamically set remaining columns
for col in range(3, ws.max_column + 1):
    col_letter = openpyxl.utils.get_column_letter(col)
    length = max(len(str(cell.value)) for cell in ws[col_letter]) * MAGIC_BUFFER
    if length > MAX_WIDTH:
        length = MAX_WIDTH
    dim_holder[col_letter] = openpyxl.worksheet.dimensions.ColumnDimension(
        ws, index=col_letter, width=length
    )

ws.column_dimensions = dim_holder

# Save workbook
wb.save(OUT)
