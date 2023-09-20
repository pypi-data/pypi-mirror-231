import openpyxl as xl


def new_sheet(wb, title):
    wb.create_sheet(title)


def cell_font(worksheet, cell, font):
    worksheet[cell].font = font


def cell_color(worksheet, cell, color):
    worksheet[cell].fill = color


def list_entry(worksheet, coulum, list):
    for items in list:
        row = worksheet.max_row + 1
        worksheet.cell(row, coulum).value = items


def cell_entry(worksheet, cell, value):
    worksheet[cell].value = value


def merge_cell(worksheet, cell1, cell2):
    worksheet.merge_cells(cell1 + ':' + cell2)


def sheet_protect(worksheet, password):
    worksheet.protection.sheet = True
    worksheet.protection.password = password


def sheet_unprotect(worksheet):
    worksheet.protection.sheet = False
