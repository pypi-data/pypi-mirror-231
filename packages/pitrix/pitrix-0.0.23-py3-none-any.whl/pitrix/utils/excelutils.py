#!/usr/bin/python3
# -*- coding: utf-8 -*-

import openpyxl
from openpyxl.styles import Font


class ExcelUtils:

    def __init__(self, file_path, sheet_name=None):
        self.file_path = file_path
        self.sheet_name = sheet_name
        try:
            self.workbook = openpyxl.load_workbook(self.file_path)
            if self.sheet_name:
                self.sheet = self.workbook[self.sheet_name]
            else:
                self.sheet = self.workbook.active
        except Exception as e:
            print(f"打开excel异常,{e}")

        self.font_color = Font(color=None)
        self.font_style = Font(name=None)

        self.font_style_dict = {
            "microsoft_accor_black": "微软雅黑",
            "regular_script": "楷体",
            "song_typeface": "宋体",
        }

        self.rgb_dict = {
            "red": "FF0000",
            "green": "FF008B00",
            "blcak": "000000"
        }

    def __enter__(self):
        try:
            self.workbook = openpyxl.load_workbook(self.file_path)
            self.sheet = self.workbook[self.sheet_name]
        except Exception as e:
            print(e)
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        if self.file_path:
            self.save_excel()

    def read_excel(self):
        """
        根据sheet_name读取excel
        :param sheet_name:
        :return:
        """
        all_rows = []

        max_row = self.sheet.max_row
        max_column = self.sheet.max_column

        key = []
        for line in range(1, max_column + 1):
            key.append(self.sheet.cell(row=1, column=line).value)  # 表格的标题

        for i in range(1, max_row + 1):
            tmpList = []
            if i < 2:
                continue
            for j in range(1, max_column + 1):
                tmpList.append(self.sheet.cell(row=i, column=j).value)
            all_rows.append(tmpList)

        rows_dict = []
        for value in all_rows:
            lis = dict(zip(key, value))
            rows_dict.append(lis)
        return rows_dict

    def write_excel(self, row=None, column=None, content=None, color='blcak'):
        """
        根据excel的sheet name 写入内容
        :param row: 写入的某一行
        :param column: 写入的某一列
        :param content:写入的内容
        :param content:写入文本的背景颜色
        :return:
        """
        self.sheet.cell(row, column).font = Font(color=self.rgb_dict[color])
        self.sheet.cell(row, column).value = content

    def save_excel(self):
        self.workbook.save(self.file_path)
