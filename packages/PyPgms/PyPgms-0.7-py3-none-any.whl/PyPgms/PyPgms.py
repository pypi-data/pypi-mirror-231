import textwrap
import os
import requests

# Program 1a
def program1a():
    program_code = textwrap.dedent('''
        print("Enter marks for")
        m1 = int(input("First test: "))
        m2 = int(input("Second test: "))
        m3 = int(input("Third test: "))

        if m1 > m2:
            if m2 > m3:
                total = m1 + m2
            else:
                total = m1 + m3
        elif m1 > m3:
            total = m1 + m2
        else:
            total = m2 + m3

        avg = total / 2
        print("The average of the best two test marks is:", avg)
    ''')

    # print(program_code)
    file_name = 'program1a.py'
    with open(file_name, 'w') as file:
        file.write(program_code)
    
    os.startfile('program1a.py')

    # print('Successfully created program1a.py')



# Program 1b
def program1b():
    program_code = textwrap.dedent('''
try:
    num = int(input("Enter any number:"))
    temp = num
    reverse = 0
    while temp > 0:
        remainder = temp % 10
        reverse = (reverse * 10) + remainder
        temp = temp // 10


    if num == reverse:
        print('Palindrome')
    else:
        print("Not Palindrome")
except:
    print("Invalid Number")
num = int(input("enter any number"))
temp = num

my_list = []
while temp != 0:
    my_list.append(temp%10)
    temp = temp // 10 

freq = {}
for item in my_list:
    if item in freq:
        freq[item] += 1
    else:
        freq[item] = 1
for item,value in freq.items():
    print(f"{item} exists {value} times")

    ''')

    # print(program_code)
    file_name = 'program1b.py'
    with open(file_name, 'w') as file:
        file.write(program_code)
        
    os.startfile('program1b.py')

    # print('Successfully created program1b.py')
    
    
# Program 2a
def program2a():
    program_code = textwrap.dedent('''
def fibonacci(nterms):
    n1, n2 = 0, 1
    count = 0
    if nterms == 1:
        print("Fibonacci sequence upto",nterms,":")
        print(n1)
    else:
        print("Fibonacci sequence:")
        while count < nterms:
            print(n1)
            nth = n1 + n2
            n1 = n2
            n2 = nth
            count += 1
nterms = int(input("How many terms to print in Fibonacci series? "))
if nterms <= 0:
    print("Please enter a positive integer")
else:
    fibonacci(nterms)

    ''')

    # print(program_code)
    file_name = 'program2a.py'
    with open(file_name, 'w') as file:
        file.write(program_code)
        
    os.startfile('program2a.py')

    # print('SSuccessfully created program2a.py')
    


# Program 2b
def program2b():
    program_code = textwrap.dedent(r'''
def BinToDec(bnum):
    dnum = 0
    i = 1
    while bnum!=0:
        rem = bnum%10
        dnum = dnum + (rem*i)
        i = i*2
        bnum = int(bnum/10)
    return dnum
def OctToHex(onum):
    chk = i = dnum = 0

    while onum!=0:
        rem = onum % 10
        if rem>7:
            chk = 1
            break
        dnum = dnum + (rem * (8 ** i))
        i = i+1
        onum = int(onum / 10)
    if chk == 0:
        hnum = ""
        while dnum != 0:
            rem = dnum % 16
            if rem < 10:
                rem = rem + 48
            else:
                rem = rem + 55
            rem = chr(rem)
            hnum = hnum + rem
            dnum = int(dnum / 16)
        hnum = hnum[::-1]

        print("\nEquivalent Hexadecimal Value =", hnum)
    else:
        print("\nInvalid Input!")

print("Enter your choice")
ch = int(input("1 for Binary to Decimal\n2 for Octal to Hexadecimal\n"))
if ch == 1:
    print("Enter the Binary Number: ", end="")
    bnum = int(input())
    print(bnum)

    dnum = BinToDec(bnum)
    print("\nEquivalent Decimal Value = ", dnum)
elif ch == 2:
    print("Enter the Octal Number: ")
    octnum = int(input())

    OctToHex(octnum)
else:
    print("Enter valid choice between 1 and 2")

    ''')

    # print(program_code)
    file_name = 'program2b.py'
    with open(file_name, 'w') as file:
        file.write(program_code)
        
    os.startfile('program2b.py')

    # print('Successfully created program2b.py')
    
    
    
# Program 3a
def program3a():
    program_code = textwrap.dedent('''
s = input("Enter a sentence: ")
w, d, u, l = 0, 0, 0, 0
l_w = s.split()
w = len(l_w)
for c in s:
    if c.isdigit():
        d = d + 1
    elif c.isupper():
        u = u + 1
    elif c.islower():
        l = l + 1
print ("No of Words: ", w)
print ("No of Digits: ", d)
print ("No of Uppercase letters: ", u)
print ("No of Lowercase letters: ", l)
    ''')

    # print(program_code)
    file_name = 'program3a.py'
    with open(file_name, 'w') as file:
        file.write(program_code)
        
    os.startfile('program3a.py')

    # print('Successfully created program3a.py')
    
    
    
# Program 3b
def program3b():
    program_code = textwrap.dedent('''
from difflib import SequenceMatcher
def is_string_similar(s1: str, s2: str):
    return SequenceMatcher(a=s1, b=s2).ratio()
print("Enter strings to be compared")
str1 = input("Enter first string: ")
str2 = input("Enter second string: ")
similar_ratio = is_string_similar(str1, str2)
print(similar_ratio)
    ''')

    # print(program_code)
    file_name = 'program3b.py'
    with open(file_name, 'w') as file:
        file.write(program_code)
        
    os.startfile('program3b.py')

    # print('Successfully created program3b.py')
    
    
    
# Program 4a
def program4a():
    program_code = textwrap.dedent('''
def insertionSort(array):
    for i in range(1, len(array)):
        next = array[i]
        prev = i - 1
        while prev >= 0 and next < array[prev]:
            array[prev + 1] = array[prev]
            prev = prev - 1
        array[prev + 1] = next

def mergeSort(arr):
    if len(arr) > 1:
        mid = len(arr) // 2
        L = arr[:mid]
        R = arr[mid:]
        mergeSort(L)
        mergeSort(R)
        i=j=k=0
        while i < len(L) and j < len(R):
            if L[i] <= R[j]:
                arr[k] = L[i]
                i += 1
            else:
                arr[k] = R[j]
                j += 1
            k += 1
        while i < len(L):
            arr[k] = L[i]
            i += 1
            k += 1
        while j < len(R):
            arr[k] = R[j]
            j += 1
            k+=1

def printList(arr):
    for i in range(len(arr)):
        print(arr[i], end =" ")
    print()

n = int(input("Enter number of list elements to be sorted "))
print("Enter list elements ")

data= []
for i in range(0,n):
    data.append(int(input()))
print()

print(data)
insertionSort(data)
print('Sorted Array in Ascending Order using Insertion Sort:')
print(data)

n = int(input("Enter number of list elements to be sorted "))
print("Enter list elements ")

data= []
for i in range(0,n):
    data.append(int(input()))
print()

print(data)
mergeSort(data)
print('Sorted Array in Ascending Order using Merge Sort:')
print(data)

    ''')

    # print(program_code)
    file_name = 'program4a.py'
    with open(file_name, 'w') as file:
        file.write(program_code)
        
    os.startfile('program4a.py')

    # print('Successfully created program4a.py')
    
    
    
# Program 4b
def program4b():
    program_code = textwrap.dedent('''
def roman_to_int(roman_num):
    roman_dict = {'I': 1, 'V': 5, 'X': 10, 'L': 510, 'C': 100, 'D': 500, 'M': 1000}
    result = 0
    prev_value = 0
    for c in roman_num[::-1]:
        curr_value = roman_dict[c]
        if curr_value < prev_value:
            result -= curr_value
        else:
            result += curr_value
        prev_value = curr_value
    return result

roman_number = input("Enter a Roman numeral: ")
integer_value = roman_to_int(roman_number)
print("The integer value is: ",integer_value)

    ''')

    # print(program_code)
    file_name = 'program4b.py'
    with open(file_name, 'w') as file:
        file.write(program_code)
        
    os.startfile('program4b.py')

    # print('Successfully created program4b.py')
    
    
    
# Program 5a
def program5a():
    program_code = textwrap.dedent('''
import re
def isPhoneWithoutReg(text):
    if len(text) != 12:
        return False
    for i in range(0, 3):
        if not text[i].isdecimal():
            return False
    if text[3] != '-':
        return False
    for i in range(4, 7):
        if not text[i].isdecimal():
            return False
    if text[7] != '-':
        return False
    for i in range(8, 12):
        if not text[i].isdecimal():
            return False
    return True
print("------------------------------")
print('Is 415-555-4242 a phone number?')
print(isPhoneWithoutReg('415-555-4242'))
print('Is Moshi moshi a phone number?')
print(isPhoneWithoutReg('Moshi moshi'))

def isPhoneReg(text):
    reg_ph = re.compile(r"\d{3}-\d{3}-\d{4}")
    try:
        mo = reg_ph.search(text)
        return mo.group()
    except:
        return False
print(isPhoneReg('Is 41-555-4242 a phone number?'))
print(isPhoneReg('Is 415-555-4242 a phone number?'))
    ''')

    # print(program_code)
    file_name = 'program5a.py'
    with open(file_name, 'w') as file:
        file.write(program_code)
        
    os.startfile('program5a.py')

    # print('Successfully created program5a.py')
    
    
    
# Program 5b
def program5b():
    program_code = textwrap.dedent(r'''
import re
phone_regex = re.compile(r'\+\d{2}\d{10}')
email_regex = re.compile(r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b')
def search_file(filename):
    with open(filename, 'r') as file:
        data = file.read()
        phone_numbers = phone_regex.findall(data)
        email_addresses = email_regex.findall(data)
        print("Phone numbers found:", phone_numbers)
        print("Email addresses found:", email_addresses)
search_file('sample.txt')
    ''')

    # print(program_code)
    file_name = 'program5b.py'
    with open(file_name, 'w') as file:
        file.write(program_code)
    
    text_file = 'sample.txt'
    with open(text_file,'w') as text:
        text.write('abcd abcd abcd +911234567890 sample@email.com 123 123 123')
        
    os.startfile('program5b.py')

    # print('Successfully created program5b.py and sample.txt')
    
    
    
# Program 6a
def program6a():
    program_code = textwrap.dedent('''
import os.path
import sys
fname = input("Enter the filename : ")
if not os.path.isfile(fname):
    print("File", fname, "doesn't exists")
    sys.exit(0)
infile = open(fname, "r")
lineList = infile.readlines()
for i in range(1):
    print(i+1,":",lineList[i])
word = input("Enter a word : ")
cnt = 0
for line in lineList:
    cnt += line.count(word)
print("The word", word, "appears", cnt, "times in the file")
    ''')

    # print(program_code)
    file_name = 'program6a.py'
    with open(file_name, 'w') as file:
        file.write(program_code)

    text_file = 'sample.txt'
    with open(text_file,'w') as text:
        text.write('abcd abcd abcd +911234567890 sample@email.com 123 123 123')
        
    os.startfile('program6a.py')

    # print('Successfully created program6a.py and sample.txt')
    
    
    
# Program 6b
def program6b():
    program_code = textwrap.dedent('''
import os
import sys
import pathlib 
import zipfile
dirName = input("Enter Directory name that you want to backup : ")
if not os.path.isdir(dirName):
    print("Directory", dirName, "doesn't exists") 
    sys.exit(0)
    
curDirectory = pathlib.Path(dirName)
with zipfile.ZipFile("myZip.zip", mode="w") as archive:
    for file_path in curDirectory.rglob("*"):
        archive.write(file_path, arcname=file_path.relative_to(curDirectory))
if os.path.isfile("myZip.zip"):
    print("Archive", "myZip.zip", "created successfully") 
else:
    print("Error in creating zip archive")


    ''')

    # print(program_code)
    file_name = 'program6b.py'
    with open(file_name, 'w') as file:
        file.write(program_code)
        
    os.startfile('program6b.py')
        
    # print('Successfully created program6b.py')
    if os.path.exists('Folder'):
        print('Folder with name Folder already exists')
    else:
        os.mkdir('Folder')
        
        # print('Folder with name Folder created successfully')
    
    
    
# Program 7a
def program7a():
    program_code = textwrap.dedent('''
import math
class Shape:
    def area(self):
        pass
class Triangle(Shape):
    def __init__(self, base, height):
        self.base = base
        self.height = height
    def area(self):
        return 0.5 * self.base * self.height
class Circle(Shape):
    def __init__(self, radius):
        self.radius = radius
    def area(self):
        return math.pi * self.radius ** 2
class Rectangle(Shape):
    def __init__(self, width, height):
        self.width = width
        self.height = height
    def area(self):
        return self.width * self.height
triangle = Triangle(5, 8)
circle = Circle(3)
rectangle = Rectangle(4, 6)
print("Area of Triangle:", triangle.area())
print("Area of Circle:", circle.area())
print("Area of Rectangle:", rectangle.area())
    ''')

    # print(program_code)
    file_name = 'program7a.py'
    with open(file_name, 'w') as file:
        file.write(program_code)
        
    os.startfile('program7a.py')

    # print('Successfully created program7a.py')
    
    
    
# Program 7b
def program7b():
    program_code = textwrap.dedent(r'''
class Employee:
    def __init__(self, name, employee_id, department, salary):
        self.name = name
        self.employee_id = employee_id
        self.department = department
        self.salary = salary
    def update_salary(self, new_salary):
        self.salary = new_salary
    def __str__(self):
        return f"Name: {self.name}\nEmployee ID:{self.employee_id}\nDepartment: {self.department}\nSalary: {self.salary}\n"
employees = [
Employee("John Doe", 1001, "Engineering", 5000),
Employee("Jane Smith", 1002, "HR", 4000),
Employee("Michael Johnson", 1003, "Finance", 6000),
Employee("Emily Davis", 1004, "Engineering", 5500)
]
department = "Engineering"
new_salary = 6000
for employee in employees:
    if employee.department == department:
        employee.update_salary(new_salary)
for employee in employees:
    print(employee)
print()

    ''')

    # print(program_code)
    file_name = 'program7b.py'
    with open(file_name, 'w') as file:
        file.write(program_code)
        
    os.startfile('program7b.py')

    # print('Successfully created program7b.py')
    
    
    
# Program 8
def program8():
    program_code = textwrap.dedent('''
class Palindrome:
    def check_palindrome(self, input_value):
        pass
class StringPalindrome(Palindrome):
    def check_palindrome(self, input_value):
        input_value = str(input_value)
        reversed_value = input_value[::-1]
        return input_value == reversed_value
class IntegerPalindrome(Palindrome):
    def check_palindrome(self, input_value):
        original_value = input_value
        reversed_value = 0
        while input_value > 0:
            digit = input_value % 10
            reversed_value = reversed_value * 10 + digit
            input_value = input_value // 10
        return original_value == reversed_value
input_string = input("enter the string ")
input_integer = int(input("enter the integer "))
string_palindrome = StringPalindrome()
integer_palindrome = IntegerPalindrome()
print("String Palindrome:", string_palindrome.check_palindrome(input_string))
print("Integer Palindrome:", integer_palindrome.check_palindrome(input_integer))
    ''')

    # print(program_code)
    file_name = 'program8.py'
    with open(file_name, 'w') as file:
        file.write(program_code)
        
    os.startfile('program8.py')

    # print('Successfully created program8.py')
    
    
    
# Program 9a
def program9a():
    program_code = textwrap.dedent('''
import requests 
import os 
from bs4 import BeautifulSoup 
url = 'https://xkcd.com/1/'
if not os.path.exists('xkcd_comics'): 
    os.makedirs('xkcd_comics') 
while True: 
    res = requests.get(url) 
    res.raise_for_status() 
    soup = BeautifulSoup(res.text, 'html.parser') 
    comic_elem = soup.select('#comic img') 
    if comic_elem == []: 
        print('Could not find comic image.') 
    else: 
        comic_url = 'https:' + comic_elem[0].get('src') 
        print(f'Downloading {comic_url}...') 
        res = requests.get(comic_url) 
        res.raise_for_status() 
        image_file = open(os.path.join('xkcd_comics', os.path.basename(comic_url)), 'wb') 
        for chunk in res.iter_content(100000): 
            image_file.write(chunk) 
        image_file.close() 
    prev_link = soup.select('a[rel="prev"]')[0] 
    if not prev_link: 
        break 
    url = 'https://xkcd.com' + prev_link.get('href') 

print('All comics downloaded.')

    ''')

    # print(program_code)
    file_name = 'program9a.py'
    with open(file_name, 'w') as file:
        file.write(program_code)
        
    os.startfile('program9a.py')

    # print('Successfully created program9a.py')
    
    
    
# Program 9b
def program9b():
    program_code = textwrap.dedent('''
from openpyxl import Workbook
from openpyxl.styles import Font

wb = Workbook()
sheet = wb.active

sheet.title = "Language"
wb.create_sheet(title="Capital")

lang = ["Kannada", "Telugu", "Tamil"]
state = ["Karnataka", "Telangana", "Tamil Nadu"]
capital = ["Bengaluru", "Hyderabad", "Chennai"]
code = ['KA', 'TS', 'TN']

sheet.cell(row=1, column=1).value = "State"
sheet.cell(row=1, column=2).value = "Language"
sheet.cell(row=1, column=3).value = "Code"

ft = Font(bold=True)

for row in sheet["A1:C1"]:
    for cell in row:
        cell.font = ft
for i in range(2, 5):
    sheet.cell(row=i, column=1).value = state[i - 2]
    sheet.cell(row=i, column=2).value = lang[i - 2]
    sheet.cell(row=i, column=3).value = code[i - 2]

wb.save("demo.xlsx")

sheet = wb["Capital"]
sheet.cell(row=1, column=1).value = "State"
sheet.cell(row=1, column=2).value = "Capital"
sheet.cell(row=1, column=3).value = "Code"
ft = Font(bold=True)

for row in sheet["A1:C1"]:
    for cell in row:
        cell.font = ft
for i in range(2, 5):
    sheet.cell(row=i, column=1).value = state[i - 2]
    sheet.cell(row=i, column=2).value = capital[i - 2]
    sheet.cell(row=i, column=3).value = code[i - 2]
wb.save("demo.xlsx")
srchCode = input("Enter state code for finding capital ")
for i in range(2, 5):
    data = sheet.cell(row=i, column=3).value
    if data == srchCode:
        print("Corresponding capital for code", srchCode, "is", sheet.cell(row=i, column=2).value)
sheet = wb["Language"]
srchCode = input("Enter state code for finding language ")
for i in range(2, 5):
    data = sheet.cell(row=i, column=3).value
    if data == srchCode:
        print("Corresponding language for code", srchCode, "is", sheet.cell(row=i,column=2).value)
wb.close()
    ''')

    # print(program_code)
    file_name = 'program9b.py'
    with open(file_name, 'w') as file:
        file.write(program_code)
        
    os.startfile('program9b.py')

    # print('Successfully created program9b.py')
    
    
    
# Program 10a
def program10a():
    program_code = textwrap.dedent('''
from PyPDF2 import PdfWriter, PdfReader
num = int(input("Enter page number you want combine from multiple documents "))
pdf1 = open('sample1.pdf', 'rb')
pdf2 = open('sample2.pdf', 'rb')
pdf_writer = PdfWriter()
pdf1_reader = PdfReader(pdf1)
page = pdf1_reader.pages[num - 1]
pdf_writer.add_page(page)
pdf2_reader = PdfReader(pdf2)
page = pdf2_reader.pages[num - 1]
pdf_writer.add_page(page)
with open('output.pdf', 'wb') as output:
    pdf_writer.write(output)

    ''')

    # print(program_code)
    file_name = 'program10a.py'
    with open(file_name, 'w') as file:
        file.write(program_code)
        
    os.startfile('program10a.py')

    # print('Successfully created program10a.py')
    try:
        url = 'https://css4.pub/2015/usenix/example.pdf'
        response = requests.get(url)
        response.raise_for_status()
        pdf1 = 'sample1.pdf'
        with open(pdf1, "wb") as file:
            file.write(response.content)
        # print('sample1')
    except:
        print('Error sample1')
    
    try:
        url = 'https://css4.pub/2017/newsletter/drylab.pdf'
        response = requests.get(url)
        response.raise_for_status()
        pdf1 = 'sample2.pdf'
        with open(pdf1, "wb") as file:
            file.write(response.content)
        # print('sample2')
    except:
        print('Error sample2')
    
    
    
# Program 10b
def program10b():
    program_code = textwrap.dedent('''
import json
with open('weather_data.json') as f:
    data = json.load(f)
current_temp = data['main']['temp']
humidity = data['main']['humidity']
weather_desc = data['weather'][0]['description']
print(f"Current temperature: {current_temp}C")
print(f"Humidity: {humidity}%")
print(f"Weather description: {weather_desc}")

    ''')
    
    # print(program_code)
    file_name = 'program10b.py'
    with open(file_name, 'w') as file:
        file.write(program_code)
        
    os.startfile('program10b.py')

    # print('Successfully created program10b.py')
    

# def program10bjson():
    json_file = textwrap.dedent('''
    {
    "coord": {
    "lon": -73.99,
    "lat": 40.73
    },
    "weather": [
    {
    "id": 800,
    "main": "Clear",
    "description": "clear sky",
    "icon": "01d"
    }
    ],
    "base": "stations",
    "main": {
    "temp": 15.45,
    "feels_like": 12.74,
    "temp_min": 14.44,
    "temp_max": 16.11,
    "pressure": 1017,
    "humidity": 64
    },
    "visibility": 10000,
    "wind": {
    "speed": 4.63,
    "deg": 180
    },
    "clouds": {
    "all": 1
    },
    "dt": 1617979985,
    "sys": {
    "type": 1,
    "id": 5141,
    "country": "US",
    "sunrise": 1617951158,
    "sunset": 1618000213
    },
    "timezone": -14400,
    "id": 5128581,
    "name": "New York",
    "cod": 200
    }
        ''')

    # print(program_code)
    file_name = 'weather_data.json'
    with open(file_name, 'w') as file:
        file.write(json_file)

    # print('Successfully created weather_data.json')
    
    
    
#Create all programs
# def all():
#     program1a()
#     program1b()
#     program2a()
#     program2b()
#     program3a()
#     program3b()
#     program4a()
#     program4b()
#     program5a()
#     program5b()
#     program6a()
#     program6b()
#     program7a()
#     program7b()
#     program8()
#     program9a()
#     program9b()
#     program10a()
#     program10b()
